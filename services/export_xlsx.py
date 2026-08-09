from __future__ import annotations
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from utils.money import to_decimal_money

SOURCE_LABELS = {
    "ru": {
        "text": "Текст",
        "voice": "Голос",
        "ocr": "Фото / OCR",
        "reminder": "Напоминание",
        "import": "Импорт",
        "miniapp": "Telegram Mini App",
        "api": "API",
        "telegram": "Текст",
    },
    "en": {
        "text": "Text",
        "voice": "Voice",
        "ocr": "Photo / OCR",
        "reminder": "Reminder",
        "import": "Import",
        "miniapp": "Telegram Mini App",
        "api": "API",
        "telegram": "Text",
    },
}

FALLBACK_COMMENTS = {"from telegram", "telegram", "from group", "from image"}


def export_comment(value: str | None) -> str:
    text = (value or "").strip()
    return "" if text.casefold() in FALLBACK_COMMENTS else text


def export_source_label(source: str | None, locale: str = "ru") -> str:
    lang = "en" if (locale or "").startswith("en") else "ru"
    raw = (source or "text").strip().lower()
    return SOURCE_LABELS[lang].get(raw, raw or SOURCE_LABELS[lang]["text"])


def _auto_width(ws):
    for col in ws.columns:
        m = 0
        letter = col[0].column_letter
        for c in col:
            m = max(m, len(str(c.value or "")))
        ws.column_dimensions[letter].width = min(max(10, m + 2), 40)


def _row_currency(row: dict, fallback_currency: str | None) -> str:
    return str(row.get("currency") or fallback_currency or "RUB").strip().upper()


def _normalized_rows(rows: list[dict], fallback_currency: str | None) -> list[dict]:
    normalized = []
    for row in rows:
        item = dict(row)
        item["amount"] = to_decimal_money(item.get("amount") or 0)
        item["currency"] = _row_currency(item, fallback_currency)
        normalized.append(item)
    return normalized


def _money_format_cells(ws, column: str, first_row: int = 2) -> None:
    for i in range(first_row, ws.max_row + 1):
        ws[f"{column}{i}"].number_format = '# ##0.00'


def build_export_xlsx(path: str, rows: list[dict], dfrom: date, dto: date, locale: str = "ru", fallback_currency: str | None = None):
    rows = _normalized_rows(rows, fallback_currency)
    wb = Workbook()
    ws = wb.active
    ws.title = "Итоги"
    ws["A1"] = "КопиPaste — экспорт операций"
    ws["A2"] = f"Период: {dfrom.strftime('%d.%m.%Y')} — {dto.strftime('%d.%m.%Y')}"
    ws.append([])
    ws.append(["Валюта", "Расходы", "Доходы", "Результат", "Операций"])
    currency_totals = defaultdict(lambda: {"expense": Decimal("0.00"), "income": Decimal("0.00"), "count": 0})
    for row in rows:
        bucket = currency_totals[row["currency"]]
        if row["type"] == "Расходы":
            bucket["expense"] += row["amount"]
        elif row["type"] == "Доходы":
            bucket["income"] += row["amount"]
        bucket["count"] += 1
    for currency, totals in sorted(currency_totals.items()):
        ws.append([currency, totals["expense"], totals["income"], totals["income"] - totals["expense"], totals["count"]])
    for c in ws[4]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9E1F2")
    _money_format_cells(ws, "B", 5)
    _money_format_cells(ws, "C", 5)
    _money_format_cells(ws, "D", 5)

    top_start = ws.max_row + 3
    ws[f"A{top_start}"] = "Топ категорий расходов:"
    ws[f"E{top_start}"] = "Топ категорий доходов:"
    grp_exp = defaultdict(lambda: Decimal("0.00"))
    grp_inc = defaultdict(lambda: Decimal("0.00"))
    for r in rows:
        key = (r["currency"], r["category"])
        if r["type"] == "Расходы":
            grp_exp[key] += r["amount"]
        elif r["type"] == "Доходы":
            grp_inc[key] += r["amount"]
    for i, ((currency, cat), sm) in enumerate(sorted(grp_exp.items(), key=lambda x: (x[0][0], -x[1], x[0][1]))[:10], start=top_start + 1):
        ws[f"A{i}"] = currency; ws[f"B{i}"] = cat; ws[f"C{i}"] = sm
    for i, ((currency, cat), sm) in enumerate(sorted(grp_inc.items(), key=lambda x: (x[0][0], -x[1], x[0][1]))[:10], start=top_start + 1):
        ws[f"E{i}"] = currency; ws[f"F{i}"] = cat; ws[f"G{i}"] = sm

    ws2 = wb.create_sheet("Операции")
    hdr = ["Дата", "Тип", "Категория", "Сумма", "Валюта", "Комментарий", "Источник", "ID"]
    ws2.append(hdr)
    for r in sorted(rows, key=lambda x: (x["op_date"], x["id"])):
        ws2.append([r["op_date"], r["type"], r["category"], float(r["amount"]), r["currency"], export_comment(r.get("comment")), export_source_label(r.get("source"), locale), r["id"]])
    for c in ws2[1]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9E1F2")
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:H{max(1, ws2.max_row)}"
    _money_format_cells(ws2, "D")

    ws3 = wb.create_sheet("По категориям")
    ws3.append(["Валюта", "Тип", "Категория", "Сумма", "Кол-во операций", "Доля от типа и валюты, %"])
    grp = defaultdict(lambda: {"sum": Decimal("0.00"), "count": 0})
    type_totals = defaultdict(lambda: Decimal("0.00"))
    for r in rows:
        k = (r["currency"], r["type"], r["category"])
        grp[k]["sum"] += r["amount"]; grp[k]["count"] += 1
        type_totals[(r["currency"], r["type"])] += r["amount"]
    for (currency, tp, cat), st in sorted(grp.items(), key=lambda x: (x[0][0], x[0][1], -x[1]["sum"])):
        total = type_totals[(currency, tp)] or 1
        ws3.append([currency, tp, cat, float(st["sum"]), st["count"], round(float(st["sum"] * Decimal("100") / total), 2)])
    _money_format_cells(ws3, "D")

    ws4 = wb.create_sheet("По месяцам")
    ws4.append(["Месяц", "Валюта", "Расходы", "Доходы", "Баланс", "Кол-во операций"])
    mgrp = defaultdict(lambda: {"e": Decimal("0.00"), "i": Decimal("0.00"), "c": 0})
    for r in rows:
        k = (r["op_date"].strftime("%Y-%m"), r["currency"])
        if r["type"] == "Расходы": mgrp[k]["e"] += r["amount"]
        if r["type"] == "Доходы": mgrp[k]["i"] += r["amount"]
        mgrp[k]["c"] += 1
    for (m, currency), st in sorted(mgrp.items()):
        ws4.append([m, currency, st["e"], st["i"], st["i"] - st["e"], st["c"]])
    _money_format_cells(ws4, "C")
    _money_format_cells(ws4, "D")
    _money_format_cells(ws4, "E")

    ws5 = wb.create_sheet("По неделям")
    ws5.append(["Неделя", "Дата начала", "Дата конца", "Валюта", "Расходы", "Доходы", "Баланс", "Кол-во операций"])
    wgrp = defaultdict(lambda: {"e": Decimal("0.00"), "i": Decimal("0.00"), "c": 0, "s": None, "e_date": None})
    for r in rows:
        d = r["op_date"]
        start = d - timedelta(days=d.weekday())
        end = start + timedelta(days=6)
        k = (start.isoformat(), r["currency"])
        wgrp[k]["s"] = start; wgrp[k]["e_date"] = end; wgrp[k]["c"] += 1
        if r["type"] == "Расходы": wgrp[k]["e"] += r["amount"]
        if r["type"] == "Доходы": wgrp[k]["i"] += r["amount"]
    for (week, currency), st in sorted(wgrp.items()):
        ws5.append([week, st["s"], st["e_date"], currency, st["e"], st["i"], st["i"] - st["e"], st["c"]])
    _money_format_cells(ws5, "E")
    _money_format_cells(ws5, "F")
    _money_format_cells(ws5, "G")

    for s in [ws, ws2, ws3, ws4, ws5]:
        _auto_width(s)
    wb.save(path)
