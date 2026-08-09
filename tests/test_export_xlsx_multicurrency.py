from datetime import date, datetime
from decimal import Decimal
import inspect

from openpyxl import load_workbook

from services.export_xlsx import build_export_xlsx


def _rows():
    return [
        {"id": 1, "op_date": date(2026, 8, 3), "type": "Расходы", "category": "Food", "amount": Decimal("10000"), "currency": "RUB", "comment": "", "source": "telegram"},
        {"id": 2, "op_date": date(2026, 8, 3), "type": "Расходы", "category": "Food", "amount": Decimal("100"), "currency": "EUR", "comment": "", "source": "miniapp"},
        {"id": 3, "op_date": date(2026, 8, 4), "type": "Доходы", "category": "Salary", "amount": Decimal("85000"), "currency": "RUB", "comment": "", "source": "telegram"},
        {"id": 4, "op_date": date(2026, 8, 4), "type": "Доходы", "category": "Bonus", "amount": Decimal("500"), "currency": "EUR", "comment": "", "source": "miniapp"},
    ]


def _values(ws):
    return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]


def test_export_xlsx_keeps_currency_totals_separate(tmp_path):
    path = tmp_path / "export.xlsx"
    build_export_xlsx(str(path), _rows(), date(2026, 8, 1), date(2026, 8, 31), fallback_currency="RUB")

    wb = load_workbook(path, data_only=True)
    summary = _values(wb["Итоги"])
    rows = [row for row in summary if row and row[0] in {"RUB", "EUR"}]

    assert any(row[:5] == ["EUR", 100, 500, 400, 2] for row in rows)
    assert any(row[:5] == ["RUB", 10000, 85000, 75000, 2] for row in rows)

    flat_numbers = [value for row in summary for value in row if isinstance(value, (int, float))]
    assert 10100 not in flat_numbers
    assert 85500 not in flat_numbers
    assert 75400 not in flat_numbers


def test_export_xlsx_operations_sheet_has_currency_and_no_ruble_format_for_eur(tmp_path):
    path = tmp_path / "export.xlsx"
    build_export_xlsx(str(path), _rows(), date(2026, 8, 1), date(2026, 8, 31), fallback_currency="RUB")

    wb = load_workbook(path, data_only=True)
    ws = wb["Операции"]

    assert [cell.value for cell in ws[1]] == ["Дата", "Тип", "Категория", "Сумма", "Валюта", "Комментарий", "Источник", "ID"]
    eur_row = next(row for row in ws.iter_rows(min_row=2) if row[4].value == "EUR")
    assert eur_row[3].value == 100
    assert "₽" not in str(eur_row[3].number_format)


def test_export_xlsx_aggregations_group_by_currency(tmp_path):
    path = tmp_path / "export.xlsx"
    build_export_xlsx(str(path), _rows(), date(2026, 8, 1), date(2026, 8, 31), fallback_currency="RUB")

    wb = load_workbook(path, data_only=True)

    categories = _values(wb["По категориям"])
    assert ["EUR", "Расходы", "Food", 100, 1, 100] in categories
    assert ["RUB", "Расходы", "Food", 10000, 1, 100] in categories

    monthly = _values(wb["По месяцам"])
    assert ["2026-08", "EUR", 100, 500, 400, 2] in monthly
    assert ["2026-08", "RUB", 10000, 85000, 75000, 2] in monthly

    weekly = _values(wb["По неделям"])
    assert ["2026-08-03", datetime(2026, 8, 3), datetime(2026, 8, 9), "EUR", 100, 500, 400, 2] in weekly
    assert ["2026-08-03", datetime(2026, 8, 3), datetime(2026, 8, 9), "RUB", 10000, 85000, 75000, 2] in weekly


def test_single_currency_export_regression(tmp_path):
    path = tmp_path / "export.xlsx"
    rows = [dict(row, currency="RUB") for row in _rows()]
    build_export_xlsx(str(path), rows, date(2026, 8, 1), date(2026, 8, 31), fallback_currency="RUB")

    wb = load_workbook(path, data_only=True)
    summary = _values(wb["Итоги"])

    assert any(row[:5] == ["RUB", 10100, 85500, 75400, 4] for row in summary)


def test_telegram_and_miniapp_export_use_shared_builder():
    from miniapp import api
    from routers import callbacks

    miniapp_source = inspect.getsource(api.MiniAppAPI.export_entry)
    callbacks_source = inspect.getsource(callbacks.callback_handler)

    assert "build_export_xlsx" in miniapp_source
    assert "fallback_currency=get_user_currency(req.user_id)" in miniapp_source
    assert "build_export_xlsx" in callbacks_source
    assert "fallback_currency=get_user_currency(cid)" in callbacks_source
